# -*- coding: utf-8 -*-
"""
Created on Thu Jan  9 16:38:56 2020

@author: pratms
"""
import time
start = time.time()
import bs4,requests
import openpyxl
path="C:\\Users\\pratms\\OneDrive - Capgemini\\Python Scripts\\excel_doc.xlsx"

wb_obj=openpyxl.load_workbook(path)
sheet_obj=wb_obj.active
m_row=sheet_obj.max_row

wb_objex = openpyxl.Workbook()
sheet_objex = wb_objex.active

finalResult=[]
bookdict={}
bookcount=1
    
url="https://medium.com/world-literature/creating-the-ultimate-list-100-books-to-read-before-you-die-45f1b722b2e5"
res=requests.get(url,verify=False)
res.raise_for_status()
soup=bs4.BeautifulSoup(res.text,'lxml')
slno=soup.find_all("a",attrs={"class":"fi cn hx hy hz ia"})
bookAndAuthors=soup.find_all("strong",attrs={"class":"id ke"})

for i in range(0,len(bookAndAuthors)-4):
    if len(bookAndAuthors[i].getText())>2:
        finalResult.append(bookAndAuthors[i].getText())
        print(bookAndAuthors[i].getText())

for i in finalResult:
    print("book "+str(bookcount)+i)
    print(len(finalResult[bookcount-1]))
    bookdict.update({bookcount:{'name':i.split(' by ')[0],'author':i.split(' by ')[1]}})
    bookcount+=1

for bn,ba in bookdict.items():
    print("book num",bn)
    booknum=bn
    for key in ba:
        print(key+":",ba[key])
        bookname=ba['name']
        bookauthor=ba['author']
    f = open(r"C:\Users\pratms\pythonprojects\webscrapping\webscrapping_medium_top_10\top100BokksByMedium.txt", "a")
    f.write(str(booknum)+'\t'+str(bookname)+'\t'+bookauthor+'\n')
    f.close()
    sheet_objex.cell(row=1,column=1).value='Sl. No'
    sheet_objex.cell(row=1,column=2).value='Book Name'
    sheet_objex.cell(row=1,column=3).value='Book Author'
    sheet_objex.cell(row=1,column=4).value='Rating'
    wb_objex.save("C:\\Users\\pratms\\OneDrive - Capgemini\\Python Scripts\\excel_doc1.xlsx")
    
###########################################################################

for i in range(1,m_row+1):
    booknumex=str(sheet_obj.cell(row=i,column=1).value)
    booknameex=str(sheet_obj.cell(row=i,column=2).value)
    bookauthorx=str(sheet_obj.cell(row=i,column=3).value)
    print(booknumex,booknameex,"by",bookauthorx)     

    
end = time.time()
print("time taken by program is:"+str(end - start)) 